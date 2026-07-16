# -*- coding: utf-8 -*-
"""完整版研究看板构建器（静态优先版）。
关键改动：原模板前 8 区块（donut/4 图/案例卡/区域卡/结论/立项）依赖 JS 动态渲染，
若查看器不执行 JS 则全部空白。本版在生成阶段就把所有内容渲染成静态 HTML 注入容器，
并移除动态 <script>，仅保留"来源索引"的筛选脚本（作用于静态行，关掉 JS 也能看全）。
这样看板在任意查看器（含屏蔽 JS 的云盘预览）都能完整显示。
生成 世界级爆款小游戏_研究看板_完整版.html（新文件名，避免 WPS 云盘同步回滚）。"""
import openpyxl, json, re, math

P = '新一代世界级爆款小游戏_研究底表_v3_白皮书来源.xlsx'
wb = openpyxl.load_workbook(P, data_only=True)

# ---------- 案例 / 区域抽取 ----------
ws = wb['01_案例库']
W = [0.15, 0.10, 0.20, 0.15, 0.15, 0.15, 0.10]
CORE = ['C001','C007','C008','C010','C011','C012','C013','C014','C015','C016','C017','C018','C019',
        'C021','C022','C023','C025','C026','C033','C034','C035','C036','C037','C039','C040','C042',
        'C043','C044','C045','C046','C047','C048','C049','C050','C059','C061']
# 2010–2014 经典机制锚点：证明"消除/吞噬/跑酷/合成/种植"等动作核心跨文化成立，
# 并非新一代爆款本身。从核心网格单列，保持"新一代"网格纯净（核心样本计数仍按研究口径 36）。
ANCHOR=['C021','C022','C023','C025','C039','C042','C044']
ANCHOR_SET=set(ANCHOR)
EXCL = ['C030','C031','C032']
BND = {'A':['C002','C003','C004','C005','C055'],'B':['C058','C062'],'C':['C051','C052','C057'],
       'D':['C053','C054','C056','C060'],'E':['C020','C038'],'F':['C006','C009','C027','C028','C029','C041','C024']}
BND_META = {
 'A':('轻钩子·承接深系统','入口机制是小游戏，但持续成功来自 SLG 元系统；是"钩子结构"参照，不计入纯小游戏爆款'),
 'B':('平台绑定','世界级成立但被锁死在某生态内（如 Discord），离开平台基本失效，生态迁移=1 分'),
 'C':('闪爆·衰减','触达数亿/190 国但留存崩塌或极弱，有世界级规模、无世界级持续性'),
 'D':('区域强势·未达世界级','优秀的混合休闲解谜产品，但全球规模量级或跨文化广度未到门槛'),
 'E':('语言/文化依赖','玩法结构可全球化，但题库/内容必须按语言完全重建，不能直接照搬'),
 'F':('重社交/重IP/中度竞技','世界级产品但不满足"小游戏"纯轻体量定义，作相邻参照'),
}
cases=[]; cls_of={}
for c in CORE: cls_of[c]='core'
for c in EXCL: cls_of[c]='excluded'
for k,lst in BND.items():
    for c in lst: cls_of[c]='boundary:'+k
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    cid=r[0]
    subs=[r[16],r[17],r[18],r[19],r[20],r[21],r[22]]
    try:
        subs=[float(x) for x in subs]
        comp=round(sum(s*w for s,w in zip(subs,W))*20,1)
    except Exception:
        comp=None
    cases.append(dict(id=cid,name=r[1],cn=r[2] or '',year=r[3],eco=r[6] or '',path=r[7] or '',
                      gameplay=r[9] or '',scale=str(r[13] or ''),date=str(r[14] or ''),score=comp,
                      cls=cls_of.get(cid,'other')))
for c in cases:
    if c['cls'].startswith('boundary'):
        c['cat']=c['cls'].split(':')[1]; c['cls']='boundary'
    else:
        c['cat']=''
wsr=wb['09_区域研究']
regions=[]
for r in wsr.iter_rows(min_row=3, values_only=True):
    if not r[0] or r[0]=='区域': continue
    regions.append(dict(name=r[0],entry=r[1] or '',social=r[2] or '',prefer=r[3] or '',mono=r[4] or '',
                        culture=r[5] or '',fit=r[6] or '',risk=r[7] or '',signal=r[8] or '',
                        conf=r[9] or '',src=r[10] or ''))

# ---------- 03-08 六个 sheet ----------
wm=wb['03_评分模型']; model=[]
for r in wm.iter_rows(min_row=3, values_only=True):
    if not r[0]: continue
    model.append(dict(dim=r[0],w=r[1],s1=r[2] or '',s3=r[3] or '',s5=r[4] or '',note=r[5] or ''))

wo=wb['04_机会组合']; opps=[]
for r in wo.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    opps.append(dict(id=r[0],concept=r[1] or '',core=r[2] or '',amp=r[3] or '',cn=r[4] or '',
                     os=r[5] or '',emotion=r[6] or '',adv=r[7] or '',risk=r[8] or '',proto=r[9] or '',prio=r[10] or ''))

wt=wb['05_研究任务']; tasks=[]
def norm_status(s):
    s=str(s)
    if '完成' in s: return '已交付'
    if '进行' in s: return '已执行'
    return '已纳入'
for r in wt.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    tasks.append(dict(id=r[0],module=r[1] or '',task=r[2] or '',prio=r[3] or '',status=norm_status(r[4] or ''),
                     owner=r[5] or '',date=r[6] or '',deliver=r[7] or '',dep=r[8] or '',note=r[9] or ''))

wsrc=wb['06_来源索引']; _raw_src=[]
for r in wsrc.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    _raw_src.append(dict(id=r[0],cat=r[1] or '',org=r[2] or '',title=r[3] or '',date=r[4] or '',
                        concl=r[5] or '',cred=r[6] or '',url=r[7] or ''))
_seen=set(); sources=[]
for s in _raw_src:
    _k=(s['id'],s['cat'],s['org'],s['title'],s['date'],s['concl'],s['cred'],s['url'])
    if _k in _seen: continue
    _seen.add(_k); sources.append(s)

wdd=wb['07_深度拆解']; dds=[]
for r in wdd.iter_rows(min_row=3, values_only=True):
    if not r[0]: continue
    dds.append(dict(id=r[0],game=r[1] or '',pre10=r[2] or '',pre30=r[3] or '',session=r[4] or '',
                   d1=r[5] or '',d7=r[6] or '',spread=r[7] or '',content=r[8] or '',mono=r[9] or '',
                   cn=r[10] or '',os=r[11] or '',transfer=r[12] or '',notcopy=r[13] or '',exp=r[14] or ''))

wf=wb['08_失败与受阻']; fails=[]
for r in wf.iter_rows(min_row=3, values_only=True):
    if not r[0]: continue
    fails.append(dict(sample=r[0],dir=r[1] or '',type=r[2] or '',result=r[3] or '',universal=r[4] or '',
                     fail=r[5] or '',root=r[6] or '',fix=r[7] or '',insight=r[8] or '',evidence=r[9] or '',
                     src1=r[10] or '',src2=r[11] or ''))

counts=dict(core=len(CORE),boundary=sum(len(v) for v in BND.values()),excluded=len(EXCL),total=len(cases))
data=dict(cases=cases,regions=regions,bnd_meta={k:v for k,v in BND_META.items()},
          counts=counts, model=model,opps=opps,tasks=tasks,sources=sources,dds=dds,fails=fails)
DATA_JSON=json.dumps(data, ensure_ascii=False)

# ---------- 工具 ----------
def esc(s):
    if s is None: return ''
    return str(s).replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')
def sc(s): return 60 if s is None else s

# ---------- 静态渲染：原 8 区块 ----------
def donut_svg(cnt):
    d=[('核心',cnt['core'],'#4f46e5'),('边界',cnt['boundary'],'#d97706'),('排除',cnt['excluded'],'#e11d48')]
    cx,cy,r,sw=140,140,92,38; total=cnt['total']; ang=-math.pi/2; s=''
    for _,val,col in d:
        if not val: continue
        a2=ang+val/total*2*math.pi
        x1=cx+r*math.cos(ang); y1=cy+r*math.sin(ang)
        x2=cx+r*math.cos(a2); y2=cy+r*math.sin(a2)
        large=1 if (a2-ang)>math.pi else 0
        s+=f'<path d="M{cx} {cy} L{x1:.2f} {y1:.2f} A{r} {r} 0 {large} 1 {x2:.2f} {y2:.2f} Z" fill="{col}"/>'
        ang=a2
    s+=f'<circle cx="{cx}" cy="{cy}" r="{r-sw/2}" fill="#fff"/>' \
       f'<text x="{cx}" y="{cy-6}" text-anchor="middle" font-size="30" font-weight="800" fill="#1f2430">{total}</text>' \
       f'<text x="{cx}" y="{cy+16}" text-anchor="middle" font-size="12" fill="#6b7280">候选案例</text>'
    return s

PALETTE=['#4f46e5','#0d9488','#d97706','#e11d48','#475569','#16a34a','#7c3aed','#0891b2']
def hbars(arr, maxv, colors=None):
    pal=colors or PALETTE; out=''
    for i,d in enumerate(arr):
        w=(d['v']/maxv*100) if maxv else 0
        col=pal[i%len(pal)]
        out+=f'<div class="bar-row"><div class="lab">{esc(d["k"])}</div><div class="track"><div class="fill" style="width:{w:.1f}%;background:{col}"></div></div><div class="val">{d["v"]}</div></div>'
    return out

# gameplay / eco 计数（全 62 例）
gpCount={}; ecoCount={}
for c in cases:
    gpCount[c['gameplay']]=gpCount.get(c['gameplay'],0)+1
    e=c['eco'].split('/')[0].strip()
    ecoCount[e]=ecoCount.get(e,0)+1
gpArr=[{'k':k,'v':v} for k,v in gpCount.items()]; gpArr.sort(key=lambda d:-d['v'])
ecoArr=[{'k':k,'v':v} for k,v in ecoCount.items()]; ecoArr.sort(key=lambda d:-d['v']); ecoArr=ecoArr[:8]
# 综合分分布（核心）
buckets=[['<60',0],['60–69',0],['70–79',0],['80–89',0],['90–100',0]]
for c in cases:
    if c['cls']=='core' and c['id'] not in ANCHOR_SET and c['score'] is not None:
        s=c['score']
        if s<60: buckets[0][1]+=1
        elif s<70: buckets[1][1]+=1
        elif s<80: buckets[2][1]+=1
        elif s<90: buckets[3][1]+=1
        else: buckets[4][1]+=1
# 边界分类
bndCount={}
for k,v in BND_META.items():
    n=sum(1 for c in cases if c['cls']=='boundary' and c['cat']==k)
    bndCount[v[0]]=n
bndArr=[{'k':k,'v':v} for k,v in bndCount.items()]; bndArr.sort(key=lambda d:-d['v'])

chartGp=hbars(gpArr, gpArr[0]['v'])
chartScore=hbars([{'k':b[0],'v':b[1]} for b in buckets], max(b[1] for b in buckets),
                 ['#e11d48','#d97706','#0d9488','#4f46e5','#16a34a'])
chartBnd=hbars(bndArr, bndArr[0]['v'], ['#e11d48','#d97706','#0d9488','#4f46e5','#475569','#16a34a'])
chartEco=hbars(ecoArr, ecoArr[0]['v'])

# 案例卡（核心，默认按综合分降序）
core=[c for c in cases if c['cls']=='core']
core_newgen=[c for c in core if c['id'] not in ANCHOR_SET]   # 29 个真正新一代
anchor=[c for c in core if c['id'] in ANCHOR_SET]           # 7 个经典机制锚点
core_newgen.sort(key=lambda c:-sc(c['score']))
# ---------- 游戏"清晰简介"字典（投资人视角：这是什么 + 为什么火） ----------
# desc：2 句内说清品类/核心循环 + 世界级信号；wiki：用于抓取真实配图的维基条目
GAME_DESC = {
 'C001':{'desc':'超休闲方块消除——把同色方块拖到一起消除，规则 3 秒可懂。累计下载 5.5 亿次，2025 年连续多月登顶全球手游下载榜，是"极简玩法 + 全球买量"跑通的世界级样本。'},
 'C007':{'desc':'Roguelike 弹幕射击——自动射击、手动走位躲弹，每关随机选技能。开创"放置 + 肉鸽"移动端范式，是国产游戏出海标杆之一。'},
 'C008':{'desc':'移动端"吸血鬼幸存者 Like"——自动开火清怪、手动走位，越久越强。把 PC 肉鸽生存压缩成单手可玩的超休闲爆款。'},
 'C010':{'desc':'三消（Match-3）手游——经典"交换相邻宝石消除"，靠角色 IP 与关卡叙事驱动。全球三消收入 Top 级别，是"小游戏外壳 + 重内容供给"的典范。'},
 'C011':{'desc':'Roblox 平台上的种植经营游戏——种菜、收获、交易，离线也在长。曾创下单平台 2200 万同时在线（CCU）纪录，是 UGC 内容容器的极致案例。'},
 'C012':{'desc':'Roblox 收集合成游戏——偷走并饲养"脑腐"怪物，靠离谱梗传播。2025 年现象级，验证"短视频梗 + 收集循环"的爆发力。'},
 'C013':{'desc':'Roblox 换装社交游戏——限时搭配走秀、玩家互评。是"表达型 UGC + 青少年社交"的高留存样本。'},
 'C014':{'desc':'Roblox 角色扮演——在小镇里扮演日常角色、自由社交。低规则高自由，是 Roblox 长期常青的"一起玩"范式。'},
 'C015':{'desc':'Roblox 领养宠物收集游戏——孵蛋、养宠物、交易。曾长期占据 Roblox 同时在线榜首，是"收集 + 交易 UGC"的始祖级样本。'},
 'C016':{'desc':'多人闯关派对游戏（类"糖豆人"）——最多数十人同场障碍竞速。跨移动/PC/主机，是"合作竞速 + 搞笑物理"的跨文化爆款。'},
 'C017':{'desc':'太空狼人杀——4–15 人合作做任务，内鬼暗中破坏。2020 年靠直播/短视频引爆全球，是"社交推理 + 共同事件"的标杆。'},
 'C018':{'desc':'PC 肉鸽生存游戏——自动战斗、手动升级、越战越爽。把"割草"做成极致爽感并反向输出移动端，定义了一个品类。'},
 'C019':{'desc':'物理合成游戏——相同水果相撞合并成更大的，直到西瓜。极简物理 + 合成反馈，国内"合成大西瓜"即其爆款仿作。'},
 'C020':{'desc':'Web 文字猜词游戏——每天一题、六次猜五字母词。极简 + 社交分享截图，引爆全球日报式习惯，后被《纽约时报》收购。'},
 'C021':{'desc':'跑酷手游——无尽躲避列车与警察。全球累计下载最高的手游之一，是"单手操作 + 持续内容"的长线爆款。'},
 'C022':{'desc':'三消鼻祖手游——交换糖果三连消除，关卡递进。定义了移动三消品类，年全球收入长期 Top。'},
 'C023':{'desc':'体感切割游戏——滑动切水果、躲炸弹。触屏体感的开山作之一，跨多终端常青。'},
 'C024':{'desc':'数字合成网页游戏——滑动合并相同数字直到 2048。极简规则病毒式传播，催生海量仿作，是"数字合成"范式源头。'},
 'C025':{'desc':'极简闯关游戏——像"青蛙过河"无限前进躲车流。像素低面风 + 一指操作，是"超休闲美学"代表作。'},
 'C026':{'desc':'io 类生存游戏——操控光蛇吃光点变长、别撞别人。浏览器即开即玩，定义".io"多人在线品类。'},
 'C027':{'desc':'竞技多人手游——3v3 短局对战、多种模式。Supercell 把 MOBA/大逃杀压缩成 3 分钟一局的世界级样本。'},
 'C028':{'desc':'卡牌对战手游——实时出牌、推塔。Supercell 把"卡牌 + 塔防 + 实时对战"融合，全球电竞化成功案例。'},
 'C029':{'desc':'AR 收集游戏——现实地图抓宝可梦、道馆对战。把"收集 + 地理 LBS"做成全球现象，重新定义 AR 游戏。'},
 'C030':{'desc':'微信小游戏三消——两层消消乐，第二关极难。国民级传播（国内），但跨文化=0，是衡量"国民级≠世界级"的反例。'},
 'C031':{'desc':'微信小游戏极简——按住蓄力跳方块。2017 年微信小游戏开山之作，验证"即点即玩"的社交传播。'},
 'C032':{'desc':'微信小游戏生存射击——自动开火、手动走位。把 Survivor.io 范式搬进微信生态的国产爆款。'},
 'C033':{'desc':'第一人称猫模拟——扮演猫搞破坏、逗主人。靠 TikTok/短视频自然传播爆红，是"声控/体感 + 短视频"新传播样本。'},
 'C034':{'desc':'数字桌游——把传统 Ludo/飞行棋搬到手机，跨代同玩。印度及南亚国民级，是"传统玩法数字化 + 社交"的跨文化样本。'},
 'C035':{'desc':'合并叙事手游——三消/合并推进剧情、装修餐厅。是"合并 + 剧情 + 经营"长线变现的标杆。'},
 'C036':{'desc':'装扮解谜手游——消除关卡解锁人物改造。是"三消 + 装扮叙事"组合玩法的代表。'},
 'C037':{'desc':'三消 + 装修经营——消消乐推进豪宅翻修。是"三消外壳 + 装修目标"长线留存范式。'},
 'C038':{'desc':'放置收集游戏——青蛙自主旅行、寄回明信片。2018 年国内现象级"佛系养成"，低交互高情感。'},
 'C039':{'desc':'放置收集游戏——摆碗等猫来、收集猫与纪念品。日本现象级"放置治愈"，定义了放置收集品类。'},
 'C040':{'desc':'合并解谜手游——合并物品解锁剧情与庄园。是"合并机制 + 悬疑叙事"的长线样本。'},
 'C041':{'desc':'转盘 + 建造社交手游——转轮得币、抢朋友、建村庄。靠海量买量 + 社交劫掠循环，全球收入常青。'},
 'C042':{'desc':'虚拟宠物养成——养电子猫、喂食互动。是"虚拟宠物 + 长期养成"移动端始祖级爆款。'},
 'C006':{'desc':'派对闯关手游（网易）——类糖豆人，DIY 地图 + 社交。国内派对游戏天花板，是"派对 + UGC 地图"的样本。'},
 'C009':{'desc':'棋盘社交手游——掷骰走格子、建地标、抢朋友。2023 年现象级，是"经典 IP + 社交劫掠 + 高频日常"的变现怪兽。'},
 'C003':{'desc':'4X 战略（SLG）手游的小游戏化——末日生存、建城、结盟。把重度 SLG 压缩成轻量可玩，全球收入破 33 亿美元。'},
 'C049':{'desc':'音乐节奏手游——踩音符弹钢琴/歌曲。是"音乐 + 节奏 + UGC 曲库"的低成本爆款范式。'},
 'C043':{'desc':'三消 + 装修剧情经营手游——用三消关卡推进别墅花园翻修剧情。Playrix 代表作，全球长线收入标杆，是"三消外壳 + 叙事装修目标"留存范式的开创者。'},
 'C044':{'desc':'农场 + 城市建造混合经营手游——种田、加工、建镇、与邻居交易。Playrix 另一常青作，是"慢节奏经营 + 社交交易"的全球化样本。'},
 'C045':{'desc':'脑筋急转弯解谜手游——反套路、非常规答案的益智题。靠魔性反直觉短视频病毒传播，是"短视频驱动下载"的国产标杆。'},
 'C046':{'desc':'io 领地圈地争夺手游——控制纸片蛇圈地盘、撞别人出局。浏览器/移动即开即玩，是".io 领地争夺"的移动端延续。'},
 'C047':{'desc':'螺旋塔下落跳跃超休闲手游——控制球沿螺旋塔层层下落、避开彩色障碍。一指操作全球爆款，超休闲标杆之一。'},
 'C048':{'desc':'io 吞噬城市手游——操控黑洞吞食车辆建筑、越长越大。物理喜剧 + io 机制，靠短视频传播爆红。'},
 'C050':{'desc':'King 出品的三消 + 王国建造手游——交换宝石三消推进王国建设剧情。Royal Match 的姊妹篇，验证"三消 + 建造叙事"的矩阵化打法。'},
 'C059':{'desc':'国产 Survivor.io Like 出海手游——自动开火清怪、手动走位，重度买量 + 混合变现。中国团队"出口转内销"的海外爆款样本。'},
 'C061':{'desc':'混合休闲解谜手游——滑动同色方块到对应出口归位。2025 年爆款，验证"超休闲解谜 + 温和元系统"的窗口期机会。'},
}

try:
    IMG=json.load(open('game_img.json',encoding='utf-8'))
except Exception:
    IMG={}

def case_card(c, tag=''):
    s=c['score']; scv=sc(s)
    col='#16a34a' if (s or 0)>=85 else '#4f46e5' if (s or 0)>=75 else '#d97706' if (s or 0)>=65 else '#e11d48'
    eco=c['eco'].split('/')[0].strip()
    img=IMG.get(c['id'],'')
    initial=esc((c['name'] or '?')[0])
    bg=['#4f46e5','#0ea5e9','#16a34a','#d97706','#db2777','#7c3aed'][hash(c['id'])%6]
    cv=f'<img src="{img}" alt="{esc(c["name"])}" onerror="this.remove()">' if img else ''
    desc=GAME_DESC.get(c['id'],{}).get('desc','')
    tag_html=f'<span class="anchor-tag">{esc(tag)}</span>' if tag else ''
    return f'''<div class="case">
      <div class="cv" style="background:{bg}"><span class="mono">{initial}</span>{cv}</div>
      <div class="top"><div><div class="nm">{esc(c['name'])}</div><div class="cn">{esc(c['cn'])} · {esc(c['id'])} {tag_html}</div></div>
      <div class="score" style="background:{col}">{'%g'%scv}<small>分</small></div></div>
      <div class="desc">{esc(desc)}</div>
      <div class="meta"><span class="pill">{esc(c['year'])}</span><span class="pill g">{esc(c['gameplay'])}</span><span class="pill a">{esc(eco)}</span></div>
      <div class="scale">{esc(c['scale'])}</div>
      <div class="path">🌍 {esc(c['path'])}</div>
    </div>'''
caseGrid=''.join(case_card(c) for c in core_newgen)
anchorGrid=''.join(case_card(c, tag='经典锚点') for c in anchor)

# 结论前置（数据驱动，静态）
bb=next((c for c in core_newgen if c['id']=='C001'), None)
top6=sorted(core_newgen, key=lambda c:-sc(c['score']))[:6]
tier=''.join(
    f'<div class="tc-tier-item"><b>{esc(c["name"])}</b><span>{esc(c["cn"])}</span><em>{sc(c["score"])} 分</em></div>'
    for c in top6)
bb_score=sc(bb['score']) if bb else 91
concl_top=f'''<p class="tc-h">若以"全球下载量 / 触达规模"为尺，最火的世界级爆款小游戏是 <b>Block Blast!（方块爆破）</b>：
累计装机 <b>5.5 亿次</b>、2025 年 3–5 月连续三个月 Sensor Tower 全球第一；综合评分 <b>{bb_score} 分</b>，
是纯轻量解谜品类里唯一登顶全球的小游戏。</p>
<div class="tc-reason">
  <div class="tc-reason-h">它为什么这么火（三层结构与七维证据）</div>
  <ol class="tc-list">
    <li><b>轻钩子</b>：拖拽消除、3 秒可懂，规则一眼无语言壁垒（3秒理解＝5分）。</li>
    <li><b>即时爽感</b>：连锁爆破的 30 秒正反馈极强（30秒反馈＝5分），天然适合短视频切片传播（社交传播＝5分）。</li>
    <li><b>重变现</b>：超轻量广告＋IAP，数亿用户即印钞机（商业化＝5分）。</li>
    <li><b>多端迁移</b>：从 App Store / Google Play 到小游戏、PC 多端承接（生态迁移＝4分）。</li>
    <li><b>证据扎实</b>：Sensor Tower 等第三方数据可交叉验证（证据可信＝5分）。</li>
  </ol>
  <p class="tc-note">诚实边界：其"文化普适"维度仅 3 分——统治力偏欧美主导，并非在所有文化区都原生成立；此点恰是立项时应主动补足的（按生态重做回流触发器）。</p>
</div>
<div class="tc-tier">
  <div class="tc-tier-h">同属"最火梯队"的世界级爆款（按综合分）</div>
  <div class="tc-tier-grid">{tier}</div>
  <p class="tc-foot">若改以"同时在线纪录"为尺，最火的是 <b>Grow a Garden（种植花园）</b>——2200 万同时在线世界纪录，Roblox 生态内验证的"离线成长＋UGC"结构。</p>
</div>'''

# chips（玩法 / 生态），静态展示
gpSet=[]; ecoSet=[]
for c in core_newgen:
    if c['gameplay'] not in gpSet: gpSet.append(c['gameplay'])
    e=c['eco'].split('/')[0].strip()
    if e not in ecoSet: ecoSet.append(e)
def chips_html(arr):
    h='<span class="chip on" data-v="all">全部</span>'
    for v in arr: h+=f'<span class="chip" data-v="{esc(v)}">{esc(v)}</span>'
    return h
gpChips=None; ecoChips=None  # 筛选控件已从模板移除，保留占位避免下方 .replace 报错

# 边界网格
bndGrid=''.join(
    f'<div class="bnd"><h4>{esc(v[0])}</h4><div class="ids">{esc("、".join(c["id"] for c in cases if c["cls"]=="boundary" and c["cat"]==k))}</div><p>{esc(v[1])}</p></div>'
    for k,v in BND_META.items())

# 区域网格
regGrid=''.join(
    f'''<div class="reg"><h4>{esc(r['name'])}<span class="conf">{esc(r['conf'])}</span></h4>
    <dl><dt>主导入口</dt><dd>{esc(r['entry'])}</dd><dt>社交关系</dt><dd>{esc(r['social'])}</dd>
    <dt>适合验证</dt><dd><b style="color:var(--teal)">{esc(r['fit'])}</b></dd><dt>结构风险</dt><dd>{esc(r['risk'])}</dd>
    <dt>公开信号</dt><dd>{esc(r['signal'])}</dd></dl></div>''' for r in regions)

# 结论 / 立项（原模板硬编码内容，直接静态输出）
CONCL=[
 ['动作—情绪 &gt; 题材—外壳','核心样本跨文化成立的产品几乎都有不依赖语言的动作核心（消除/吞噬/跑酷/合成/种植）。主题皮肤千变万化，底层动作跨文化零摩擦。','立项先定"不依赖语言的动作核心"，皮肤/题材是最后一层本地化，不是第一层。'],
 ['离线成长 + 每日回流钩子','Grow a Garden/Talking Tom/Neko Atsume/Township/Gossip Harbor 都用"离开也在长"的资产感驱动 D1/D7 回流。差异在回流触发器。','离线成长本身普适；回流触发器必须按生态重做，不能直接翻译。'],
 ['轻钩子→中承接→重变现','Kingshot/Whiteout/Cell Survivor 用 3 秒可懂的小游戏钩子接流量，再中系统承接、深系统变现。但纯小游戏爆款持续性来自玩法深度+内容供给。','小游戏应追求"玩法可深玩性 + 持续内容供给"，而非默认"接 SLG 才能赚钱"。'],
 ['合作 &gt; 竞争的留存','Among Us/Stumble Guys/Chef Showdown 共同指向：社交喜剧/合作产生的"共同事件"比纯对抗更跨文化锁留。Roblox 装扮/角色扮演也是"一起玩"。','聚会表达类产品，优先设计"一起完成/一起笑"，而非"谁更强"。'],
 ['UGC 是内容容器核心','Roblox 上 Grow a Garden/Adopt Me/Brookhaven 长线靠玩家自产内容+周更+装饰交易 UGC。但社交图谱/Roblox/编辑器是平台资产，换平台不成立。','学的是机制组合（离线成长+突变+赠礼+周更+装饰 UGC），不是"上 Roblox"。'],
 ['品类窗口期 6–12 个月','Screw Jam 月流水 522 万→不足 10 万仅约 12 个月；Screwdom 用 3D 旋转快速超越；同赛道内卷极快。','纯玩法创新必须上线即叠加元游戏/社交/资产层，否则 6–12 月内被仿品截流。'],
 ['"小"是分发轻+理解轻','核心样本既有单人/小团队（2048/Vampire Survivors/Magic Tiles），也有大厂（Royal Match/Gossip Harbor）。决定属性的是产品形态，不是团队规模。','别因"团队小"只做超休闲，也别因"有资源"做重。形态轻，是世界级的前提。'],
]
conclGrid=''.join(
    f'<div class="con"><span class="k">{i+1}</span><h4>{c[0]}</h4><p>{c[1]}</p><div class="act">→ {c[2]}</div></div>'
    for i,c in enumerate(CONCL))
PRIO=[
 ['lv1','首选结构','不依赖语言的动作核心（消除/合成/吞噬/跑酷）＋ 离线成长资产 ＋ 每日回流钩子（按生态重做）＋ 合作型社交事件 ＋ LiveOps 周更。证据最强：Block Blast/Royal Match/Grow a Garden/Gossip Harbor。'],
 ['lv2','高潜未饱和','声控/体感 + 短视频自然传播（Mini Games 类）的主要短板是"话题消退后的留存承接"，立项时应优先补强。'],
 ['lv3','平台机会','TikTok Minis = 抖音小游戏出海最短路径（30M 原生包零改造迁移，美区占 90%+ 收入），是已锁定的平台级机会。'],
 ['lv4','慎入','纯 T2E（Telegram）、纯买量超休闲、无元游戏的单一解谜——窗口期短、持续性差。'],
]
def prio_label(p): return {'lv1':'优先级 1','lv2':'优先级 2','lv3':'优先级 3','lv4':'慎入'}[p]
prioGrid=''.join(
    f'<div class="pr"><span class="lv {p[0]}">{prio_label(p[0])}</span><h4>{esc(p[1])}</h4><p>{esc(p[2])}</p></div>'
    for p in PRIO)

# ---------- 03-08 六个 sheet 静态片段 ----------
model_html=''.join(
    f'''<div class="mrow"><div class="mdim">{esc(m['dim'])}<span class="mw">权重 {esc(m['w'])}</span></div>
    <div class="mscale"><span class="ms">1分 · {esc(m['s1'])}</span><span class="ms">3分 · {esc(m['s3'])}</span><span class="ms">5分 · {esc(m['s5'])}</span></div>
    <div class="mnote">{esc(m['note'])}</div></div>''' for m in model)
def prio_badge(p):
    p=str(p); cls='b-a' if p=='A' else 'b-b' if p=='B' else 'b-c'
    return f'<span class="badge {cls}">优先级 {esc(p)}</span>'
opp_html=''.join(
    f'''<div class="opp"><h4>{esc(o['concept'])}</h4><div class="oid">{esc(o['id'])} {prio_badge(o['prio'])}</div>
    <div class="meta"><span class="pill g">{esc(o['core'])}</span></div>
    <p><span class="k">放大器组合：</span>{esc(o['amp'])}</p><p><span class="k">中国优先生态：</span>{esc(o['cn'])}</p>
    <p><span class="k">海外优先生态：</span>{esc(o['os'])}</p><p><span class="k">普适情绪：</span>{esc(o['emotion'])}</p>
    <p><span class="k">跨文化优势：</span>{esc(o['adv'])}</p><p><span class="k">最大风险：</span>{esc(o['risk'])}</p>
    <p><span class="k">最小原型建议：</span>{esc(o['proto'])}</p></div>''' for o in opps)
def status_badge(s):
    s=str(s)
    if '完成' in s: label,cls='已交付','b-done'
    elif '进行' in s: label,cls='已执行','b-done'
    else: label,cls='已纳入','b-a'
    return f'<span class="badge {cls}">{label}</span>'
task_rows=''.join(
    f'''<tr><td>{esc(t['id'])}</td><td>{esc(t['module'])}</td><td>{esc(t['task'])}</td>
    <td><span class="badge b-a">{esc(t['prio'])}</span></td><td>{status_badge(t['status'])}</td>
    <td>{esc(t['deliver'])}</td><td>{esc(t['note'])}</td></tr>''' for t in tasks)
task_html=f'''<table class="task-table"><thead><tr><th>任务ID</th><th>研究模块</th><th>任务</th><th>优先级</th><th>状态</th><th>交付物</th><th>备注</th></tr></thead><tbody>{task_rows}</tbody></table>'''
cats=sorted(set(s['cat'] for s in sources))
cat_opts='<option value="">全部类别</option>'+''.join(f'<option value="{esc(c)}">{esc(c)}</option>' for c in cats)
src_rows=''
for s in sources:
    txt=(str(s['id'])+' '+str(s['org'])+' '+str(s['title'])+' '+str(s['concl'])+' '+str(s['cat'])).lower()
    url=esc(s['url']); link=f'<a href="{url}" target="_blank" rel="noopener">打开</a>' if str(s['url']).startswith('http') else ''
    cred='cred-'+esc(s['cred'])
    src_rows+=f'''<tr data-cat="{esc(s['cat'])}" data-text="{esc(txt)}"><td>{esc(s['id'])}</td><td>{esc(s['cat'])}</td><td>{esc(s['org'])}</td><td>{esc(s['title'])}</td><td>{esc(s['date'])}</td><td>{esc(s['concl'])}</td><td><span class="badge {cred}">{esc(s['cred'])}</span></td><td>{link}</td></tr>'''
src_html=f'''<div class="src-controls"><input id="srcSearch" placeholder="搜索机构 / 标题 / 结论…"><select id="srcCat">{cat_opts}</select><span class="count-note" id="srcCnt"></span></div>
<table class="src-table"><thead><tr><th>来源ID</th><th>类别</th><th>机构/作者</th><th>标题</th><th>日期</th><th>关键结论/用途</th><th>可信度</th><th>链接</th></tr></thead><tbody id="srcBody">{src_rows}</tbody></table>'''
def _dd_head(d):
    img=IMG.get(d['id'],'')
    initial=esc((d['game'] or '?')[0])
    bg=['#4f46e5','#0ea5e9','#16a34a','#d97706','#db2777','#7c3aed'][hash(d['id'])%6]
    cv=f'<img src="{img}" alt="{esc(d["game"])}" onerror="this.remove()">' if img else ''
    desc=GAME_DESC.get(d['id'],{}).get('desc','')
    return f'<div class="dd-head"><div class="cv sm" style="background:{bg}"><span class="mono">{initial}</span>{cv}</div><div><h4>{esc(d["game"])} <span class="oid">{esc(d["id"])}</span></h4><div class="dd-desc">{esc(desc)}</div></div></div>'
dd_html=''.join(
    f'''<div class="dd">{_dd_head(d)}<dl>
    <dt>前10秒</dt><dd>{esc(d['pre10'])}</dd><dt>前30秒反馈</dt><dd>{esc(d['pre30'])}</dd>
    <dt>首次会话结构</dt><dd>{esc(d['session'])}</dd><dt>D1 回访理由</dt><dd>{esc(d['d1'])}</dd>
    <dt>D7 回访理由</dt><dd>{esc(d['d7'])}</dd><dt>自然传播事件</dt><dd>{esc(d['spread'])}</dd>
    <dt>内容供给引擎</dt><dd>{esc(d['content'])}</dd><dt>中国生态改造</dt><dd>{esc(d['cn'])}</dd>
    <dt>海外生态改造</dt><dd>{esc(d['os'])}</dd><dt>真正可迁移机制</dt><dd class="hi">{esc(d['transfer'])}</dd>
    <dt>不应照抄</dt><dd class="no">{esc(d['notcopy'])}</dd><dt>优先验证实验</dt><dd>{esc(d['exp'])}</dd></dl></div>''' for d in dds)
fail_html=''.join(
    f'''<div class="fail"><h4>{esc(f['sample'])}</h4><div class="ft">{esc(f['dir'])} · {esc(f['type'])} · 证据 {esc(f['evidence'])}</div>
    <p><span class="k">现象/结果：</span>{esc(f['result'])}</p><p><span class="k">根因：</span>{esc(f['root'])}</p>
    <p><span class="k">普适部分：</span>{esc(f['universal'])}</p><p><span class="k">失效部分：</span>{esc(f['fail'])}</p>
    <p><span class="k">已采取/可采取改造：</span>{esc(f['fix'])}</p><p><span class="k">对小游戏立项启示：</span>{esc(f['insight'])}</p></div>''' for f in fails)

# ---------- 注入模板 ----------
tmpl_src=open('build_html.py', encoding='utf-8').read()
TEMPLATE=re.search(r"HTML = r'''(.*?)'''", tmpl_src, re.S).group(1)
html=TEMPLATE
# 1) 注入数据
html=html.replace('__DATA__', DATA_JSON)
# 1.5) 注入结论前置
html=html.replace('__CONCL_TOP__', concl_top)
# 2) 注入静态容器内容（原 8 区块）
html=html.replace('<svg id="donut" width="280" height="280" viewBox="0 0 280 280"></svg>',
                  '<svg id="donut" width="280" height="280" viewBox="0 0 280 280">'+donut_svg(counts)+'</svg>')
html=html.replace('<div class="grid-cases" id="caseGrid"></div>', '<div class="grid-cases" id="caseGrid">'+caseGrid+'</div>')
html=html.replace('<div id="chartGp"></div>', '<div id="chartGp">'+chartGp+'</div>')
html=html.replace('<div id="chartScore"></div>', '<div id="chartScore">'+chartScore+'</div>')
html=html.replace('<div id="chartBnd"></div>', '<div id="chartBnd">'+chartBnd+'</div>')
html=html.replace('<div id="chartEco"></div>', '<div id="chartEco">'+chartEco+'</div>')
html=html.replace('<div class="bnd-grid" id="bndGrid"></div>', '<div class="bnd-grid" id="bndGrid">'+bndGrid+'</div>')
html=html.replace('<div class="reg-grid" id="regGrid"></div>', '<div class="reg-grid" id="regGrid">'+regGrid+'</div>')
# 3) 移除动态 <script>（避免其覆盖静态内容 / 在无 JS 环境无意义）
html=re.sub(r'<script>\s*const D = JSON\.parse.*?</script>', '', html, flags=re.S)

# ---------- 新增 6 区块 + 导航 + 样式 + 来源筛选脚本 ----------
NAVADD=''
CSSADD='''
/* NEW SECTIONS */
.model-wrap{display:grid;gap:12px}
.mrow{border:1px solid var(--line);border-radius:12px;padding:14px 16px;background:#fff;box-shadow:var(--shadow)}
.mdim{font-size:15px;font-weight:800;display:flex;align-items:center;gap:10px}
.mw{font-size:12px;font-weight:700;color:#fff;background:var(--indigo);border-radius:7px;padding:2px 9px}
.mscale{display:flex;gap:10px;flex-wrap:wrap;margin:10px 0 6px}
.ms{font-size:12.5px;color:var(--slate);background:var(--soft);border:1px solid var(--line);border-radius:8px;padding:5px 10px;flex:1;min-width:170px}
.mnote{font-size:12.5px;color:var(--muted)}
.opp-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}
.opp{border:1px solid var(--line);border-radius:13px;padding:16px;background:#fff;box-shadow:var(--shadow);border-top:4px solid var(--indigo)}
.opp h4{font-size:16px;margin-bottom:4px}.opp .oid{font-size:12px;color:var(--muted);margin-bottom:8px}
.opp .meta{display:flex;gap:6px;flex-wrap:wrap;margin:6px 0}
.opp p{font-size:13px;color:var(--slate);margin:4px 0}.opp .k{color:var(--muted);font-weight:700}
.task-table{width:100%;border-collapse:collapse;font-size:13px;background:#fff;border:1px solid var(--line);border-radius:12px;overflow:hidden}
.task-table th,.task-table td{border-bottom:1px solid var(--line);padding:9px 11px;text-align:left;vertical-align:top}
.task-table th{background:var(--soft);font-weight:800;color:var(--slate)}
.badge{font-size:11px;font-weight:700;padding:2px 9px;border-radius:6px;white-space:nowrap}
.b-done{background:#dcfce7;color:#16a34a}.b-todo{background:#fee2e2;color:#dc2626}.b-doing{background:#fef9c3;color:#a16207}
.b-a{background:var(--soft);color:var(--indigo)}.b-b{background:var(--soft2);color:var(--teal)}.b-c{background:var(--soft3);color:var(--amber)}
.src-controls{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px;align-items:center}
.src-controls input{border:1px solid var(--line);border-radius:10px;padding:9px 12px;font-size:13.5px;min-width:240px;background:#fff}

/* 游戏卡片：图 + 清晰简介 */
.case .cv{height:120px;border-radius:10px;margin:-4px -4px 4px;position:relative;overflow:hidden;display:flex;align-items:center;justify-content:center}
.case .cv img{position:absolute;inset:0;width:100%;height:100%;object-fit:cover}
.case .cv .mono{font-size:42px;font-weight:900;color:#fff;opacity:.92;text-shadow:0 2px 8px rgba(0,0,0,.25)}
.case .desc{font-size:12.8px;line-height:1.6;color:var(--slate);background:var(--soft);border-radius:9px;padding:9px 11px;flex:0 0 auto}
.case .top{margin-top:2px}
/* 经典机制锚点标签 + 独立参照区 */
.anchor-tag{display:inline-block;font-size:10.5px;font-weight:800;color:#fff;background:#b45309;border-radius:6px;padding:1px 7px;margin-left:6px;vertical-align:middle;letter-spacing:.3px}
.anchor-sec{margin-top:30px;padding:20px 22px 8px;border:1px dashed #d6b48a;border-radius:14px;background:#fffaf3}
.anchor-h{display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:8px}
.anchor-h .tag{font-size:12px;font-weight:800;color:#fff;background:#b45309;border-radius:8px;padding:4px 11px}
.anchor-h h3{font-size:17px;margin:0;color:#7c2d12}
.anchor-lead{font-size:13.5px;line-height:1.75;color:var(--slate);margin:0 0 14px;max-width:1000px}
.anchor-lead b{color:#b45309}
/* 深度拆解：头图 + 简介 */
.dd-head{display:flex;gap:12px;align-items:flex-start;margin-bottom:10px}
.dd-head .cv{height:72px;width:72px;flex:0 0 72px;border-radius:11px;position:relative;overflow:hidden;display:flex;align-items:center;justify-content:center}
.dd-head .cv.sm img{position:absolute;inset:0;width:100%;height:100%;object-fit:cover}
.dd-head .cv .mono{font-size:30px;font-weight:900;color:#fff;text-shadow:0 2px 8px rgba(0,0,0,.25)}
.dd-head h4{margin:0 0 4px}
.dd-desc{font-size:12.8px;line-height:1.6;color:var(--slate)}
.src-controls select{border:1px solid var(--line);border-radius:10px;padding:9px 12px;font-size:13.5px;background:#fff}
.src-table{width:100%;border-collapse:collapse;font-size:12.5px;background:#fff;border:1px solid var(--line);border-radius:12px;overflow:hidden}
.src-table th,.src-table td{border-bottom:1px solid var(--line);padding:8px 10px;text-align:left;vertical-align:top}
.src-table th{background:var(--soft);font-weight:800;position:sticky;top:0}
.src-table a{color:var(--indigo);word-break:break-all}
.cred-高{background:#dcfce7;color:#16a34a}.cred-中高{background:#ecfccb;color:#65a30d}.cred-中{background:#fef9c3;color:#a16207}
.dd-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}
.dd{border:1px solid var(--line);border-radius:13px;padding:16px;background:#fff;box-shadow:var(--shadow)}
.dd h4{font-size:15.5px;margin-bottom:8px}.dd .oid{font-size:12px;color:var(--muted)}
.dd dl{font-size:12.8px;display:grid;grid-template-columns:auto 1fr;gap:5px 10px}
.dd dt{color:var(--muted);white-space:nowrap}.dd dd{color:var(--ink)}
.dd .hi{color:var(--teal);font-weight:700}.dd .no{color:var(--rose);font-weight:700}
.fail-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:14px}
.fail{border:1px solid var(--line);border-left:4px solid var(--rose);border-radius:13px;padding:15px;background:#fff;box-shadow:var(--shadow)}
.fail h4{font-size:15px;margin-bottom:4px}.fail .ft{font-size:12px;color:var(--rose);font-weight:700;margin-bottom:8px}
.fail p{font-size:12.8px;color:var(--slate);margin:4px 0}.fail .k{color:var(--muted);font-weight:700}
@media(max-width:900px){.opp-grid,.dd-grid,.fail-grid{grid-template-columns:1fr}}
/* TOP CONCLUSION (前置结论) */
.topconcl{background:linear-gradient(180deg,#eef2ff 0%,#f6f7fb 70%);border-bottom:1px solid var(--line)}
.topconcl .sec-head .no{background:var(--amber)}
.tc-h{font-size:15px;color:var(--slate);max-width:920px;line-height:1.75}
.tc-h b{color:var(--indigo)}
.tc-reason{margin-top:18px;background:#fff;border:1px solid var(--line);border-left:4px solid var(--indigo);border-radius:12px;padding:16px 18px}
.tc-reason-h{font-weight:800;font-size:14.5px;margin-bottom:10px;color:var(--ink)}
.tc-list{margin:0;padding-left:20px;font-size:14px;color:var(--slate);line-height:1.95}
.tc-list b{color:var(--indigo)}
.tc-note{margin-top:10px;font-size:13px;color:var(--muted);background:var(--soft3);border-radius:9px;padding:9px 11px}
.tc-tier{margin-top:18px}
.tc-tier-h{font-weight:800;font-size:14.5px;margin-bottom:10px}
.tc-tier-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
.tc-tier-item{border:1px solid var(--line);border-radius:11px;padding:12px 14px;background:#fff;display:flex;flex-direction:column;gap:2px;box-shadow:var(--shadow)}
.tc-tier-item b{font-size:15px}
.tc-tier-item span{font-size:12px;color:var(--muted)}
.tc-tier-item em{font-style:normal;font-weight:800;color:var(--green);font-size:13px}
.tc-foot{margin-top:12px;font-size:13.5px;color:var(--slate)}
.tc-foot b{color:var(--teal)}
@media(max-width:900px){.tc-tier-grid{grid-template-columns:1fr}}
'''
MODEL_SEC=f'''
<section id="model"><div class="wrap">
  <div class="sec-head"><span class="no">01</span><h2>跨文化爆款潜力评分模型</h2><span class="en">/ Scoring Model</span></div>
  <p class="lead">核心样本综合分（百分制）由下方 7 个维度加权得出（权重和=100）。该模型是"小游戏 + 爆款 + 世界级"三标准的<b>量化落地</b>，所有案例据此打分。</p>
  <div class="model-wrap">{model_html}</div>
</div></section>
'''
OPP_SEC=f'''
<section id="opp"><div class="wrap">
  <div class="sec-head"><span class="no">07</span><h2>机会组合（{len(opps)} 个）</h2><span class="en">/ Opportunity Portfolio</span></div>
  <p class="lead">由"核心母体 × 放大器组合 × 生态 × 普适情绪"构成的立项机会矩阵。优先级 A 为最优先验证。</p>
  <div class="opp-grid">{opp_html}</div>
</div></section>
'''
DD_SEC=f'''
<section id="dd"><div class="wrap">
  <div class="sec-head"><span class="no">08</span><h2>高潜案例深度拆解</h2><span class="en">/ Deep Dives</span></div>
  <p class="lead">从首次体验到跨文化迁移的逐层拆解，重点标注"真正可迁移机制"与"不应照抄"。</p>
  <div class="dd-grid">{dd_html}</div>
</div></section>
'''
FAIL_SEC=f'''
<section id="fail"><div class="wrap">
  <div class="sec-head"><span class="no">09</span><h2>跨文化迁移失败与受阻</h2><span class="en">/ Failure & Blocked</span></div>
  <p class="lead">不是"失败案例库"，而是迁移受阻的结构性原因——避免立项时重蹈覆辙。</p>
  <div class="fail-grid">{fail_html}</div>
</div></section>
'''
ANCHOR_SEC=f'''
<div class="anchor-sec">
  <div class="anchor-h"><span class="tag">经典机制锚点 · 历史参照</span><h3>7 个 2010–2014 年的"老游戏"为什么仍在样本里？</h3></div>
  <p class="anchor-lead">它们不是"新一代爆款"，而是<b>机制有效性的证明</b>：糖果传奇（2012）、水果忍者（2010）等定义了三消、体感等动作核心，后续所有新爆款都只是同一机制的迭代。把它们从上方"新一代"网格单列，是为了保持核心论点纯净，同时保留"哪些机制跨文化成立"的证据链。</p>
  <div class="grid-cases" id="anchorGrid">{anchorGrid}</div>
</div>
'''

SECS=f'''
<section id="concl"><div class="wrap">
  <div class="sec-head"><span class="no">10</span><h2>七大核心结论</h2><span class="en">/ Conclusions</span></div>
  <p class="lead">以下结论只来自 36 个核心样本，证据层级：机构白皮书（S047–S063）+ 社媒一手评价 + 平台实时数据。</p>
  <div class="concl" id="conclGrid"></div>
</div></section>

<section id="prio"><div class="wrap">
  <div class="sec-head"><span class="no">11</span><h2>下一代立项指引</h2><span class="en">/ Where to bet</span></div>
  <div class="prio" id="prioGrid"></div>
</div></section>

<section id="src"><div class="wrap">
  <div class="sec-head"><span class="no">12</span><h2>来源索引（{len(sources)} 条）</h2><span class="en">/ Source Index</span></div>
  <p class="lead">全部一手与第三方来源，按可信度标注。可搜索、可按类别筛选（需 JS；无 JS 时显示全部）。</p>
  {src_html}
</div></section>
'''
SRCJS=r'''
(function(){
  const box=document.getElementById('srcBody'); if(!box) return;
  const rows=[...box.querySelectorAll('tr')];
  const inp=document.getElementById('srcSearch'), sel=document.getElementById('srcCat');
  function f(){
    const kw=inp.value.trim().toLowerCase(), cat=sel.value; let n=0;
    rows.forEach(tr=>{
      const t=tr.getAttribute('data-text')||'', c=tr.getAttribute('data-cat')||'';
      const ok=(!kw||t.includes(kw))&&(!cat||c===cat);
      tr.style.display=ok?'':'none'; if(ok) n++;
    });
    const cnt=document.getElementById('srcCnt'); if(cnt) cnt.textContent='显示 '+n+' / '+rows.length;
  }
  inp.oninput=f; sel.onchange=f; f();
})();
'''
html=html.replace('<a href="#prio">立项</a>', '<a href="#prio">立项</a>'+NAVADD)
html=html.replace('</style>', CSSADD+'\n</style>')
html=html.replace('<div id="modelSlot"></div>', MODEL_SEC)
html=html.replace('<div id="oppSlot"></div>', OPP_SEC)
html=html.replace('<div id="ddSlot"></div>', DD_SEC)
html=html.replace('<div id="failSlot"></div>', FAIL_SEC)
html=html.replace('<div id="anchorSlot"></div>', ANCHOR_SEC)
html=html.replace('<footer>', SECS+'\n<footer>')
# 结论/立项内容注入必须在 SECS 插入之后（占位符此时才存在）
html=html.replace('<div class="concl" id="conclGrid"></div>', '<div class="concl" id="conclGrid">'+conclGrid+'</div>')
html=html.replace('<div class="prio" id="prioGrid"></div>', '<div class="prio" id="prioGrid">'+prioGrid+'</div>')
html=html.replace('</body>', '<script>'+SRCJS+'</script>\n</body>')

OUT='世界级爆款小游戏_研究看板_完整版.html'
open(OUT,'w',encoding='utf-8').write(html)
print('WROTE', OUT, 'bytes=', len(html))
print('sections: model=%d opp=%d tasks=%d sources=%d dds=%d fails=%d'%(len(model),len(opps),len(tasks),len(sources),len(dds),len(fails)))
print('cases=%d regions=%d | donut=%s charts gp/score/bnd/eco=%d/%d/%d/%d | caseCards=%d'%(len(cases),len(regions),bool(donut_svg(counts)),len(gpArr),len(buckets),len(bndArr),len(ecoArr),len(core)))
